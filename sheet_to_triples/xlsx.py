# Copyright 2020-2021 Visual Meaning Ltd
# This is free software licensed as GPL-3.0-or-later - see COPYING for terms.

"""Reading of new-format Excel files."""

import warnings

import openpyxl


def _clean(cell):
    if isinstance(cell.value, str):
        cell.value = openpyxl.utils.escape.unescape(cell.value)
    return cell


class Book:
    """Wrapper around openpyxl.Workbook to present common interface."""

    def __init__(self, book):
        self._book = book

    def __repr__(self):
        return f'{self.__class__.__name__}({self._book})'

    @classmethod
    def from_path(cls, path):
        with warnings.catch_warnings():
            warnings.filterwarnings('ignore', '.*extension is not supported')
            return cls(openpyxl.load_workbook(path, data_only=True))

    def iter_rows_in_sheet(self, sheet, *args):
        return (
            tuple(_clean(cell) for cell in row)
            for row in self._book[sheet].iter_rows()
        )
